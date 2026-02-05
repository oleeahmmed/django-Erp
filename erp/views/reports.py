"""
Report Views - Sales, Purchase, and Stock Reports
"""
from django.contrib import admin
from django.shortcuts import render
from django.views import View
from django.utils.decorators import method_decorator
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, F, Count, Q
from django.http import HttpResponse
from decimal import Decimal
from datetime import datetime

from ..models import (
    SalesOrder, SalesOrderItem, PurchaseOrder, PurchaseOrderItem,
    Product, Customer, Supplier, Category, SalesPerson
)


class SalesReportView(View):
    """Enhanced Sales Report View with Excel Export"""
    
    @method_decorator(staff_member_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        # Get filter parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        customer_id = request.GET.get('customer')
        salesperson_id = request.GET.get('salesperson')
        product_id = request.GET.get('product')
        status = request.GET.get('status')
        export_excel = request.GET.get('export') == 'excel'
        
        # Base queryset - get sales order items for detailed reporting
        order_items = SalesOrderItem.objects.select_related(
            'sales_order', 'sales_order__customer', 'sales_order__salesperson', 'product'
        ).order_by('-sales_order__order_date', '-sales_order__order_number')
        
        # Apply filters
        if from_date:
            order_items = order_items.filter(sales_order__order_date__gte=from_date)
        if to_date:
            order_items = order_items.filter(sales_order__order_date__lte=to_date)
        if customer_id:
            order_items = order_items.filter(sales_order__customer_id=customer_id)
        if salesperson_id:
            order_items = order_items.filter(sales_order__salesperson_id=salesperson_id)
        if product_id:
            order_items = order_items.filter(product_id=product_id)
        if status:
            order_items = order_items.filter(sales_order__status=status)
        
        # Excel export
        if export_excel:
            return self.export_to_excel(order_items, request)
        
        # Calculate statistics
        total_items = order_items.count()
        total_revenue = order_items.aggregate(total=Sum('line_total'))['total'] or Decimal('0.00')
        total_quantity = order_items.aggregate(total=Sum('quantity'))['total'] or Decimal('0.00')
        
        # Get unique orders count
        unique_orders = order_items.values('sales_order').distinct().count()
        
        # Get filter options
        customers = Customer.objects.filter(is_active=True).order_by('name')
        salespersons = SalesPerson.objects.filter(is_active=True).order_by('name')
        products = Product.objects.filter(is_active=True).order_by('name')
        
        context = {
            **admin.site.each_context(request),
            'title': 'Sales Report',
            'subtitle': 'Detailed sales analysis with filters',
            'order_items': order_items[:500],  # Limit for performance
            'total_items': total_items,
            'total_revenue': total_revenue,
            'total_quantity': total_quantity,
            'unique_orders': unique_orders,
            'customers': customers,
            'salespersons': salespersons,
            'products': products,
            'filters': {
                'from_date': from_date or '',
                'to_date': to_date or '',
                'customer': customer_id or '',
                'salesperson': salesperson_id or '',
                'product': product_id or '',
                'status': status or '',
            }
        }
        
        return render(request, 'admin/erp/sales_report.html', context)
    
    def export_to_excel(self, order_items, request):
        """Export sales report to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Header style
        header_fill = PatternFill(start_color="C4D82E", end_color="C4D82E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Order #', 'Order Date', 'Customer', 'Salesperson',
            'Product', 'Quantity', 'Unit Price', 'Line Total', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row, item in enumerate(order_items, 2):
            ws.cell(row=row, column=1, value=item.sales_order.order_number).border = border
            ws.cell(row=row, column=2, value=item.sales_order.order_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=3, value=item.sales_order.customer.name).border = border
            ws.cell(row=row, column=4, value=item.sales_order.salesperson.name if item.sales_order.salesperson else 'N/A').border = border
            ws.cell(row=row, column=5, value=item.product.name).border = border
            ws.cell(row=row, column=6, value=float(item.quantity)).border = border
            ws.cell(row=row, column=7, value=float(item.unit_price)).border = border
            ws.cell(row=row, column=8, value=float(item.line_total)).border = border
            ws.cell(row=row, column=9, value=item.sales_order.get_status_display()).border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response


class PurchaseReportView(View):
    """Enhanced Purchase Report View with Excel Export"""
    
    @method_decorator(staff_member_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        # Get filter parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        supplier_id = request.GET.get('supplier')
        product_id = request.GET.get('product')
        status = request.GET.get('status')
        export_excel = request.GET.get('export') == 'excel'
        
        # Base queryset - get purchase order items for detailed reporting
        order_items = PurchaseOrderItem.objects.select_related(
            'purchase_order', 'purchase_order__supplier', 'product'
        ).order_by('-purchase_order__order_date', '-purchase_order__order_number')
        
        # Apply filters
        if from_date:
            order_items = order_items.filter(purchase_order__order_date__gte=from_date)
        if to_date:
            order_items = order_items.filter(purchase_order__order_date__lte=to_date)
        if supplier_id:
            order_items = order_items.filter(purchase_order__supplier_id=supplier_id)
        if product_id:
            order_items = order_items.filter(product_id=product_id)
        if status:
            order_items = order_items.filter(purchase_order__status=status)
        
        # Excel export
        if export_excel:
            return self.export_to_excel(order_items, request)
        
        # Calculate statistics
        total_items = order_items.count()
        total_amount = order_items.aggregate(total=Sum('line_total'))['total'] or Decimal('0.00')
        total_quantity = order_items.aggregate(total=Sum('quantity'))['total'] or Decimal('0.00')
        
        # Get unique orders count
        unique_orders = order_items.values('purchase_order').distinct().count()
        
        # Get filter options
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')
        products = Product.objects.filter(is_active=True).order_by('name')
        
        context = {
            **admin.site.each_context(request),
            'title': 'Purchase Report',
            'subtitle': 'Detailed purchase analysis with filters',
            'order_items': order_items[:500],  # Limit for performance
            'total_items': total_items,
            'total_amount': total_amount,
            'total_quantity': total_quantity,
            'unique_orders': unique_orders,
            'suppliers': suppliers,
            'products': products,
            'filters': {
                'from_date': from_date or '',
                'to_date': to_date or '',
                'supplier': supplier_id or '',
                'product': product_id or '',
                'status': status or '',
            }
        }
        
        return render(request, 'admin/erp/purchase_report.html', context)
    
    def export_to_excel(self, order_items, request):
        """Export purchase report to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Report"
        
        # Header style
        header_fill = PatternFill(start_color="C4D82E", end_color="C4D82E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Order #', 'Order Date', 'Supplier',
            'Product', 'Quantity', 'Unit Price', 'Line Total', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row, item in enumerate(order_items, 2):
            ws.cell(row=row, column=1, value=item.purchase_order.order_number).border = border
            ws.cell(row=row, column=2, value=item.purchase_order.order_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=3, value=item.purchase_order.supplier.name).border = border
            ws.cell(row=row, column=4, value=item.product.name).border = border
            ws.cell(row=row, column=5, value=float(item.quantity)).border = border
            ws.cell(row=row, column=6, value=float(item.unit_price)).border = border
            ws.cell(row=row, column=7, value=float(item.line_total)).border = border
            ws.cell(row=row, column=8, value=item.purchase_order.get_status_display()).border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=purchase_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response


class StockReportView(View):
    """Stock/Inventory Report View"""
    
    @method_decorator(staff_member_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        # Get filter parameters
        category_id = request.GET.get('category')
        stock_status = request.GET.get('stock_status')
        
        # Default queryset
        products = Product.objects.select_related('category').filter(is_active=True).order_by('name')
        
        # Apply filters
        if category_id:
            products = products.filter(category_id=category_id)
        
        if stock_status == 'low':
            products = products.filter(current_stock__lte=F('min_stock_level'))
        elif stock_status == 'out':
            products = products.filter(current_stock=0)
        elif stock_status == 'in':
            products = products.filter(current_stock__gt=F('min_stock_level'))
        
        # Calculate statistics
        total_products = products.count()
        total_stock_value = sum(
            float(p.current_stock * p.purchase_price) for p in products
        )
        low_stock_count = Product.objects.filter(
            current_stock__lte=F('min_stock_level'),
            is_active=True
        ).count()
        out_of_stock_count = Product.objects.filter(
            current_stock=0,
            is_active=True
        ).count()
        
        # Get limit from request
        limit = request.GET.get('limit', '100')
        if limit == 'all':
            limited_products = products
            showing_limited = False
        else:
            try:
                limit_int = int(limit)
                limited_products = products[:limit_int]
                showing_limited = total_products > limit_int
            except ValueError:
                limited_products = products[:100]
                showing_limited = total_products > 100
                limit = '100'
        
        categories = Category.objects.filter(is_active=True).order_by('name')
        
        context = {
            **admin.site.each_context(request),
            'title': 'Stock Report',
            'subtitle': 'View and filter inventory records',
            'products': limited_products,
            'categories': categories,
            'total_products': total_products,
            'total_stock_value': total_stock_value,
            'low_stock_count': low_stock_count,
            'out_of_stock_count': out_of_stock_count,
            'showing_limited': showing_limited,
            'current_limit': limit,
            'filters': {
                'category': category_id or '',
                'stock_status': stock_status or '',
            }
        }
        
        return render(request, 'admin/erp/stock_report.html', context)
